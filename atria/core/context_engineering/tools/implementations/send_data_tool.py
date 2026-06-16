"""send_data tool — push tabular data (CSV/XLSX) to the web UI as a chart bubble."""

from __future__ import annotations

import csv
import io
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import httpx


_MAX_BYTES = 10 * 1024 * 1024  # 10 MB
_MAX_ROWS = 10_000
_MAX_COLS = 50
_URL_TIMEOUT = 30

_ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_BOOL_SET = frozenset({"true", "false", "0", "1"})
_XLSX_MAGIC = b"PK\x03\x04"


def _err(msg: str) -> dict[str, Any]:
    """Build a uniform failure dict for the tool result."""
    return {"success": False, "error": msg, "output": None}


class SendDataHandler:
    """Handler for the send_data tool.

    Parses CSV/XLSX/XLS files (from a server-side path or http(s) URL) into
    JSON rows plus a column schema with inferred types, then broadcasts the
    payload to the web UI via ``ui_callback.on_data``.
    """

    def send(self, args: dict[str, Any], context: Any) -> dict[str, Any]:
        path = (args.get("path") or "").strip()
        url = (args.get("url") or "").strip()
        title = (args.get("title") or "").strip()
        suggestions = args.get("suggestions")

        if bool(path) == bool(url):
            return _err("Provide exactly one of 'path' or 'url'")

        if not title:
            return _err("'title' is required")

        if not isinstance(suggestions, list) or not suggestions:
            return _err("'suggestions' must be a non-empty list")

        for i, sug in enumerate(suggestions):
            if not isinstance(sug, dict):
                return _err(f"suggestion[{i}] must be an object")
            if not sug.get("chart_type") or not isinstance(sug.get("chart_type"), str):
                return _err(f"suggestion[{i}] missing 'chart_type'")
            if not sug.get("x") or not isinstance(sug.get("x"), str):
                return _err(f"suggestion[{i}] missing 'x' column")
            y = sug.get("y")
            if not isinstance(y, list) or not y or not all(isinstance(c, str) for c in y):
                return _err(f"suggestion[{i}] 'y' must be a non-empty list of strings")

        ui_callback = getattr(context, "ui_callback", None)
        if ui_callback is None or not hasattr(ui_callback, "on_data"):
            return _err("UI callback unavailable; send_data only works in the web UI")

        # Load raw bytes and decide extension
        loaded = self._load_url(url) if url else self._load_path(path)
        if not loaded.get("success"):
            return _err(loaded.get("error", "Failed to load file"))
        raw: bytes = loaded["data"]
        ext: str = loaded["ext"]

        # Parse
        try:
            if ext == "csv":
                header, data_rows = self._parse_csv(raw)
            elif ext in ("xlsx", "xls"):
                header, data_rows = self._parse_xlsx(raw)
            else:
                return _err(f"Unsupported file extension: {ext}")
        except Exception as exc:  # noqa: BLE001
            return _err(f"Failed to parse file: {exc}")

        if not data_rows:
            return _err("File contains no data rows")

        # Dedup header names
        header = self._dedup_header(header)

        warnings: list[str] = []

        # Truncate columns
        if len(header) > _MAX_COLS:
            warnings.append(
                f"Column count {len(header)} exceeded limit; truncated to first {_MAX_COLS}"
            )
            header = header[:_MAX_COLS]
            data_rows = [row[:_MAX_COLS] for row in data_rows]

        # Truncate rows
        if len(data_rows) > _MAX_ROWS:
            warnings.append(
                f"Row count {len(data_rows)} exceeded limit; truncated to first {_MAX_ROWS}"
            )
            data_rows = data_rows[:_MAX_ROWS]

        # Normalize cells to strings (or "") for inference
        n_cols = len(header)
        norm_rows: list[list[str]] = []
        for row in data_rows:
            # Pad/truncate to n_cols
            if len(row) < n_cols:
                row = list(row) + [""] * (n_cols - len(row))
            elif len(row) > n_cols:
                row = list(row)[:n_cols]
            norm_rows.append([self._cell_to_str(c) for c in row])

        # Infer column types
        col_types: list[str] = []
        for ci in range(n_cols):
            col_values = [r[ci] for r in norm_rows]
            col_types.append(self._infer_column_type(col_values))

        columns = [{"name": header[i], "type": col_types[i]} for i in range(n_cols)]

        # Build typed rows
        rows: list[dict[str, Any]] = []
        for r in norm_rows:
            obj: dict[str, Any] = {}
            for i, cell in enumerate(r):
                obj[header[i]] = self._cast_value(cell, col_types[i])
            rows.append(obj)

        # Validate suggestion column refs
        col_name_set = set(header)
        for i, sug in enumerate(suggestions):
            if sug["x"] not in col_name_set:
                return _err(
                    f"suggestion[{i}].x references unknown column '{sug['x']}'. "
                    f"Available: {sorted(col_name_set)}"
                )
            for y_col in sug["y"]:
                if y_col not in col_name_set:
                    return _err(
                        f"suggestion[{i}].y references unknown column '{y_col}'. "
                        f"Available: {sorted(col_name_set)}"
                    )

        warning_text = "; ".join(warnings) if warnings else None

        payload: dict[str, Any] = {
            "title": title,
            "columns": columns,
            "rows": rows,
            "suggestions": suggestions,
        }
        if warning_text:
            payload["warning"] = warning_text

        ui_callback.on_data(payload)

        result: dict[str, Any] = {
            "success": True,
            "output": f"Sent {len(rows)} rows × {n_cols} cols to UI",
            # data_payload is persisted on the ToolCall (DB) and used by the
            # frontend to re-render the chart bubble on session reload. The LLM
            # only sees `output` — payload bypasses the prompt entirely.
            "data_payload": payload,
        }
        if warning_text:
            result["warning"] = warning_text
        return result

    # ------------------------------------------------------------------
    # Loaders
    # ------------------------------------------------------------------

    def _load_path(self, path: str) -> dict[str, Any]:
        p = Path(path)
        if not p.is_absolute():
            return {"success": False, "error": "Path must be absolute"}
        if not p.exists() or not p.is_file():
            return {"success": False, "error": f"File not found: {path}"}
        size = p.stat().st_size
        if size > _MAX_BYTES:
            return {"success": False, "error": f"File too large ({size} bytes; max {_MAX_BYTES})"}
        ext = p.suffix.lower().lstrip(".")
        if ext not in ("csv", "xlsx", "xls"):
            return {"success": False, "error": f"Unsupported file extension: .{ext}"}
        return {"success": True, "data": p.read_bytes(), "ext": ext}

    def _load_url(self, url: str) -> dict[str, Any]:
        if not (url.startswith("http://") or url.startswith("https://")):
            return {"success": False, "error": "URL must start with http:// or https://"}

        # Best-effort extension from URL path
        url_ext = ""
        tail = url.split("?", 1)[0].rsplit("/", 1)[-1]
        if "." in tail:
            url_ext = tail.rsplit(".", 1)[-1].lower()

        content_type = ""
        buf = bytearray()
        try:
            with httpx.stream("GET", url, timeout=_URL_TIMEOUT, follow_redirects=True) as resp:
                resp.raise_for_status()
                content_type = (resp.headers.get("Content-Type") or "").lower()
                for chunk in resp.iter_bytes(chunk_size=64 * 1024):
                    buf.extend(chunk)
                    if len(buf) > _MAX_BYTES:
                        return {
                            "success": False,
                            "error": f"Remote file exceeds {_MAX_BYTES} bytes",
                        }
        except httpx.HTTPError as exc:
            return {"success": False, "error": f"Failed to fetch URL: {exc}"}
        data = bytes(buf)

        # Decide ext
        ext = url_ext if url_ext in ("csv", "xlsx", "xls") else ""
        if not ext:
            if "csv" in content_type or "text/plain" in content_type:
                ext = "csv"
            elif "spreadsheetml" in content_type or "excel" in content_type:
                ext = "xlsx"
            elif data.startswith(_XLSX_MAGIC):
                ext = "xlsx"
            else:
                # Default to CSV if the body looks like text
                try:
                    data[: min(len(data), 4096)].decode("utf-8")
                    ext = "csv"
                except UnicodeDecodeError:
                    return {
                        "success": False,
                        "error": "Could not determine file type from URL",
                    }
        return {"success": True, "data": data, "ext": ext}

    # ------------------------------------------------------------------
    # Parsers
    # ------------------------------------------------------------------

    def _parse_csv(self, raw: bytes) -> tuple[list[str], list[list[str]]]:
        # Decode with utf-8 (fallback to latin-1 to never fail)
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw.decode("latin-1")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return [], []
        header = [str(c).strip() for c in rows[0]]
        data_rows: list[list[str]] = []
        for row in rows[1:]:
            # Skip wholly-empty rows
            if not any(str(c).strip() for c in row):
                continue
            data_rows.append([str(c) for c in row])
        return header, data_rows

    def _parse_xlsx(self, raw: bytes) -> tuple[list[str], list[list[str]]]:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        header: list[str] = []
        data_rows: list[list[str]] = []
        try:
            ws = wb.active
            if ws is None:
                return [], []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    header = [str(c).strip() if c is not None else "" for c in row]
                    continue
                # Skip wholly empty
                if not any(c is not None and str(c).strip() != "" for c in row):
                    continue
                data_rows.append([c if c is not None else "" for c in row])  # type: ignore[misc]
        finally:
            wb.close()
        # Stringify cells in data_rows (datetime preserved as ISO for parser)
        out_rows: list[list[str]] = []
        for row in data_rows:
            out_rows.append([self._cell_to_str(c) for c in row])
        return header, out_rows

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _cell_to_str(cell: Any) -> str:
        if cell is None:
            return ""
        if isinstance(cell, datetime):
            return cell.isoformat()
        if isinstance(cell, bool):
            return "true" if cell else "false"
        if isinstance(cell, float):
            # Avoid trailing ".0" loss vs original — keep float repr; inference handles it.
            return repr(cell) if not cell.is_integer() else str(int(cell))
        return str(cell)

    @staticmethod
    def _dedup_header(header: list[str]) -> list[str]:
        out: list[str] = []
        seen: dict[str, int] = {}
        for h in header:
            name = h.strip() if h else ""
            if not name:
                name = "column"
            if name not in seen:
                seen[name] = 1
                out.append(name)
            else:
                seen[name] += 1
                out.append(f"{name}_{seen[name]}")
        return out

    @staticmethod
    def _try_float(s: str) -> float | None:
        if s == "":
            return None
        # Strip thousand separators
        candidate = s.replace(",", "")
        try:
            return float(candidate)
        except ValueError:
            return None

    @staticmethod
    def _try_date(s: str) -> str | None:
        if s == "":
            return None
        if _ISO_DATE_RE.match(s):
            return s
        v = s
        if v.endswith("Z"):
            v = v[:-1]
        try:
            dt = datetime.fromisoformat(v)
            return dt.isoformat()
        except ValueError:
            return None

    @classmethod
    def _infer_column_type(cls, values: list[str]) -> str:
        """Pick the narrowest type that fits every non-empty cell — single pass."""
        all_number = all_date = all_bool = True
        seen_any = False
        for v in values:
            if v == "":
                continue
            seen_any = True
            if all_number and cls._try_float(v) is None:
                all_number = False
            if all_date and cls._try_date(v) is None:
                all_date = False
            if all_bool and v.lower() not in _BOOL_SET:
                all_bool = False
            if not (all_number or all_date or all_bool):
                return "string"
        if not seen_any:
            return "string"
        if all_number:
            return "number"
        if all_date:
            return "date"
        if all_bool:
            return "bool"
        return "string"

    @classmethod
    def _cast_value(cls, cell: str, col_type: str) -> Any:
        if cell == "":
            return None
        if col_type == "number":
            v = cls._try_float(cell)
            return v if v is not None else None
        if col_type == "date":
            return cls._try_date(cell)
        if col_type == "bool":
            low = cell.lower()
            if low in ("true", "1"):
                return True
            if low in ("false", "0"):
                return False
            return None
        return cell
