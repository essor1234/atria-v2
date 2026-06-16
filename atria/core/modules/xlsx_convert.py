"""Convert uploaded Excel workbooks to CSV so module scripts/dashboards can read them.

Used by the module data-upload endpoint. Kept separate from ``store.py`` so the
optional ``openpyxl`` dependency is only imported when a conversion is requested.
"""

from __future__ import annotations

import csv
import io
import re
from typing import List, Tuple


def _slug(text: str) -> str:
    """Filesystem-safe lowercase slug for a sheet name."""
    text = (text or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "sheet"


def xlsx_to_csvs(xlsx_bytes: bytes, base_name: str) -> List[Tuple[str, bytes]]:
    """Convert an .xlsx/.xlsm workbook to one CSV per (non-empty) worksheet.

    Args:
        xlsx_bytes: Raw workbook bytes.
        base_name: Base file name (without extension) for the output CSV(s).

    Returns:
        List of ``(filename, csv_bytes)``. A single-sheet workbook yields
        ``<base_name>.csv``; a multi-sheet workbook yields
        ``<base_name>__<sheet-slug>.csv`` per sheet.

    Raises:
        Exception: Propagates openpyxl load errors (caller decides how to report).
    """
    from openpyxl import load_workbook  # imported lazily — optional dependency

    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    try:
        sheets = list(wb.worksheets)
        multi = len(sheets) > 1
        out: List[Tuple[str, bytes]] = []
        for ws in sheets:
            buf = io.StringIO(newline="")
            writer = csv.writer(buf)
            wrote_any = False
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if cell is None else cell for cell in row])
                wrote_any = True
            if not wrote_any:
                continue
            fname = f"{base_name}__{_slug(ws.title)}.csv" if multi else f"{base_name}.csv"
            out.append((fname, buf.getvalue().encode("utf-8")))
        return out
    finally:
        wb.close()
